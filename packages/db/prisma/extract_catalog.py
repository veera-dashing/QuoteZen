"""Extract every catalog/reference table from the Quote Base workbook into catalog.json.

Run:  python3 extract_catalog.py "<path to xlsx>"
Output: packages/db/prisma/data/catalog.json  (consumed by import-catalogs.ts)
"""
import json
import sys
import os
import openpyxl


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def s(v):
    if v is None:
        return None
    return str(v).strip() or None


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}

    # ── led_products: (LED 1) catalog rows (vendor in col A, model in col B) ──
    ws = wb["(LED 1)"]
    led = []
    for r in range(5, 234):
        vendor = s(ws.cell(r, 1).value)
        model = s(ws.cell(r, 2).value)
        if not model or not vendor:
            continue
        led.append({
            "vendor": vendor,
            "model": model,
            "upgradeOptions": s(ws.cell(r, 3).value),
            "mechanicalOptions": s(ws.cell(r, 4).value),
            "volumetricModifier": num(ws.cell(r, 7).value),
            "kgPerSqm": num(ws.cell(r, 8).value),
            "costPerSqmUsd": num(ws.cell(r, 9).value),
            "moduleWMm": num(ws.cell(r, 10).value),
            "moduleHMm": num(ws.cell(r, 11).value),
            "minCabinetWMm": num(ws.cell(r, 12).value),
            "minCabinetHMm": num(ws.cell(r, 13).value),
            "cabinetDepthMm": num(ws.cell(r, 14).value),
            "powerMaxW": num(ws.cell(r, 15).value),
            "powerAvgW": num(ws.cell(r, 16).value),
            "shipDepthMm": num(ws.cell(r, 17).value),
            "pixelPitchH": num(ws.cell(r, 18).value),
            "pixelPitchV": num(ws.cell(r, 19).value),
            "brightnessNits": num(ws.cell(r, 20).value),
            "serviceAccess": s(ws.cell(r, 21).value),
            "cabinetType": s(ws.cell(r, 22).value),
            "includesReceivers": s(ws.cell(r, 23).value) == "Y",
            "gobIncluded": s(ws.cell(r, 24).value) == "Y",
            "packIncluded": s(ws.cell(r, 25).value) == "Y",
            "modulePrice": num(ws.cell(r, 26).value),
        })
    # de-dup by model (the sheet repeats the selected product at top)
    seen = set()
    out["ledProducts"] = [x for x in led if not (x["model"] in seen or seen.add(x["model"]))]

    # ── display_catalog: LCDRef, category = nearest section header above ──
    ws = wb["LCDRef"]
    sections = {
        3: "Philips Indoor", 52: "Philips Commercial Chromecast", 59: "Video Wall",
        70: "Malaysia", 74: "Other Indoor", 86: "High Bright", 99: "High Bright Outdoor",
        115: "Touch", 137: "Stretch Panels", 145: "Projectors", 164: "Gobo Projectors",
        171: "Videri", 181: "All in One Screens", 194: "Mediaplayers", 204: "Peripherals",
        252: "Nexmosphere", 273: "4G Networking", 282: "Standard Brackets", 283: "Aidoru",
        301: "Selby", 311: "Specialist", 330: "Vogels Basic Brackets", 353: "Vogels Modular",
        387: "Vogels Kits", 394: "Projector Brackets", 406: "Shrouds & Culpans",
    }
    section_rows = sorted(sections)
    disp = []
    for r in range(5, 544):
        model = s(ws.cell(r, 2).value)
        if not model or model.startswith("SELECT"):
            continue
        sell = num(ws.cell(r, 11).value)
        aud = num(ws.cell(r, 7).value)
        if sell is None and aud is None:
            continue
        cat = "Other"
        for sr in section_rows:
            if sr <= r:
                cat = sections[sr]
        out_cat = cat if r < 282 else "Brackets" if cat in (
            "Standard Brackets", "Aidoru", "Selby", "Specialist", "Vogels Basic Brackets",
            "Vogels Modular", "Vogels Kits", "Projector Brackets") else cat
        disp.append({
            "category": out_cat,
            "subcategory": cat if out_cat != cat else None,
            "sizeInch": num(ws.cell(r, 1).value),
            "model": model,
            "description": s(ws.cell(r, 3).value),
            "usd": num(ws.cell(r, 4).value),
            "listAud": aud,
            "freight": num(ws.cell(r, 8).value) or 0,
            "totalCost": num(ws.cell(r, 9).value),
            "margin": num(ws.cell(r, 10).value),
            "sell": sell,
        })
    out["displayCatalog"] = disp

    # ── import_catalog: Import (Philips Q-Line) ──
    ws = wb["Import"]
    imp = []
    for r in range(1, 86):
        model = s(ws.cell(r, 4).value)
        if not model:
            continue
        imp.append({
            "brand": s(ws.cell(r, 1).value) or "PHILIPS",
            "series": s(ws.cell(r, 2).value),
            "sizeInch": s(ws.cell(r, 3).value),
            "model": model,
            "description": s(ws.cell(r, 5).value),
            "cost": num(ws.cell(r, 6).value),
            "sell": num(ws.cell(r, 7).value),
            "partNumber": s(ws.cell(r, 9).value),
        })
    out["importCatalog"] = imp

    # ── audio_products: Audio (category = section header) ──
    ws = wb["Audio"]
    audio = []
    cat = "Yamaha"
    for r in range(2, 62):
        a = s(ws.cell(r, 1).value)
        cost = num(ws.cell(r, 4).value)
        sell = num(ws.cell(r, 5).value)
        if a and cost is None and sell is None and "http" not in (a or ""):
            if a not in ("Yamaha",):
                cat = a
            continue
        name = a or s(ws.cell(r, 3).value)
        if not name or (cost is None and sell is None):
            continue
        audio.append({
            "category": cat, "name": name,
            "sourceUrl": s(ws.cell(r, 2).value), "cost": cost, "sell": sell,
        })
    out["audioProducts"] = audio

    # ── music_services: Music ──
    ws = wb["Music"]
    music = []
    for r in range(2, 25):
        name = s(ws.cell(r, 1).value)
        if not name:
            continue
        cost = num(ws.cell(r, 2).value)
        sell = num(ws.cell(r, 3).value)
        if cost is None and sell is None:
            continue
        cat = "OneMusic" if "One Music" in name or "OneMusic" in name else \
              "Soundtrack" if "SoundTrack" in name or "Option" in name else "Other"
        music.append({"category": cat, "name": name, "cost": cost, "sell": sell})
    out["musicServices"] = music

    # ── hypervsn_products: Hypervsn (A..F) ──
    ws = wb["Hypervsn"]
    hv = []
    cat = "Unit"
    cat_map = {"Wall KIts": "Wall Kit", "SmartV": "SmartV", "Rental": "Rental",
               "Standard Services": "Services", "Peripherals": "Peripheral", "Spares": "Spare"}
    for r in range(3, 61):
        name = s(ws.cell(r, 1).value)
        b = num(ws.cell(r, 2).value)
        if name and b is None and num(ws.cell(r, 3).value) is None:
            for k, v in cat_map.items():
                if name.startswith(k):
                    cat = v
            continue
        if not name:
            continue
        if b is None and num(ws.cell(r, 3).value) is None:
            continue
        c = "Cloud Licence" if "Licence" in name or "Licen" in name else cat
        hv.append({
            "category": c, "name": name,
            "sellAud": b, "resellerAud": num(ws.cell(r, 3).value),
            "sellNzd": num(ws.cell(r, 5).value), "resellerNzd": num(ws.cell(r, 6).value),
        })
    out["hypervsnProducts"] = hv

    # ── software_activities: Software Costs ──
    ws = wb["Software Costs"]
    sw = []
    for r in range(3, 16):
        act = s(ws.cell(r, 1).value)
        cost = num(ws.cell(r, 2).value)
        sell = num(ws.cell(r, 3).value)
        if not act or (cost is None and sell is None):
            continue
        sw.append({"activity": act, "cost": cost or 0, "sell": sell or 0,
                   "ratio": num(ws.cell(r, 5).value)})
    out["softwareActivities"] = sw

    # ── led_commentary: Reference Data 294-316 ──
    ws = wb["Reference Data"]
    comm = []
    svc = "Indoor Front Service"
    svc_headers = {"Indoor Front Service", "Indoor Rear Service", "Outdoor Front Service",
                   "Transparent", "Rental"}
    for r in range(295, 317):
        a = s(ws.cell(r, 1).value)
        b = s(ws.cell(r, 2).value)
        if a and not b and a in svc_headers:
            svc = a
            continue
        if a and b:
            comm.append({"serviceCategory": svc, "productCode": a, "commentary": b})
    out["ledCommentary"] = comm

    # ── international_install_rates: Reference Data 275-286 ──
    intl = []
    for r in range(275, 287):
        a = s(ws.cell(r, 1).value)
        v = num(ws.cell(r, 2).value)
        if not a or v is None:
            continue
        region = "SG" if "(SG" in a or "SG," in a or "Ambience" in a else \
                 "RSA" if "RSA" in a else "USA" if "USA" in a else "UK" if "UK" in a else None
        partner = a.split(" (")[0] if " (" in a else a.split(" ")[0]
        intl.append({"partner": partner, "region": region, "rateLabel": a, "cost": v})
    out["internationalInstallRates"] = intl

    # ── international_vat: Reference Data 289-292 ──
    vat = []
    for r in range(289, 293):
        a = s(ws.cell(r, 1).value)
        v = num(ws.cell(r, 2).value)
        if a and v is not None:
            vat.append({"region": a, "vatMultiplier": v})
    out["internationalVat"] = vat

    # ── hardware_support_components: Licence & Support rows 17-19 ──
    ws = wb["Licence & Support"]
    hsc = []
    for r, comp in [(17, "Support Per Screen"), (18, "Interactive uplift Per Screen"),
                    (19, "Site Fee")]:
        for tier, lcd_c, led_c in [("low", 3, 4), ("high", 5, 6)]:
            for st, col in [("LCD", lcd_c), ("LED", led_c)]:
                v = num(ws.cell(r, col).value)
                if v is not None:
                    hsc.append({"component": comp, "tier": tier, "screenType": st, "value": v})
    out["hardwareSupportComponents"] = hsc

    # ── installer_rates: Installer Breakdown row 2 ──
    ws = wb["Installer Breakdown"]
    inst = []
    for r in range(2, 3):
        region = s(ws.cell(r, 1).value)
        if not region:
            continue
        inst.append({
            "region": region, "location": s(ws.cell(r, 2).value),
            "installer": s(ws.cell(r, 3).value), "lcd": num(ws.cell(r, 4).value),
            "led": num(ws.cell(r, 5).value), "bracket": num(ws.cell(r, 6).value),
            "customWorks": num(ws.cell(r, 7).value), "customAdditional": num(ws.cell(r, 8).value),
            "permit": num(ws.cell(r, 9).value), "disposal": num(ws.cell(r, 10).value),
            "eveningWorks": num(ws.cell(r, 11).value), "gst": num(ws.cell(r, 12).value),
        })
    out["installerRates"] = inst

    # ── manufactured_products: Manufactured matrix (one product per config column) ──
    ws = wb["Manufactured"]
    man = []
    for c in range(4, 178):
        t = s(ws.cell(1, c).value)
        if not t:
            continue
        cost = num(ws.cell(158, c).value)   # Total Cost row
        sell = num(ws.cell(159, c).value)   # Sell Direct @33.3% row
        if cost is None and sell is None:
            continue
        bright = num(ws.cell(3, c).value)
        man.append({
            "type": t,
            "sizeInch": s(ws.cell(2, c).value),
            "brightness": int(bright) if bright is not None else None,
            "cost": cost,
            "sell": sell,
        })
    out["manufacturedProducts"] = man

    # ── international_support_rates: Licence & Support partner blocks (cols H-M) ──
    ws = wb["Licence & Support"]
    partner_map = {
        "UK (": ("UX Global", "UK"), "South Africa": ("Corporate AV", "RSA"),
        "US and Canada": ("Snider", "US"), "SG - Ambience": ("Ambience", "SG"),
        "SG - Live": ("Live Acoustics", "SG"),
    }
    intl_sup = []
    partner = region = currency = None
    for r in range(1, 60):
        h = s(ws.cell(r, 8).value)
        if not h:
            continue
        matched = next((v for k, v in partner_map.items() if h.startswith(k)), None)
        if matched:
            partner, region = matched
            currency = None
            continue
        if h.startswith("Service Per Technician"):
            currency = s(ws.cell(r, 9).value)  # GBP / ZAR / USD / SGD
            continue
        if h in ("VAT", "Markup") or not partner:
            continue
        aud = num(ws.cell(r, 10).value)        # AUD column
        if aud is None:
            continue
        intl_sup.append({
            "partner": partner, "region": region, "rateLabel": h,
            "localValue": num(ws.cell(r, 9).value), "localCurrency": currency,
            "audValue": aud, "sellValue": num(ws.cell(r, 12).value) or num(ws.cell(r, 11).value),
        })
    out["internationalSupportRates"] = intl_sup

    here = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(here, "data")
    os.makedirs(data_dir, exist_ok=True)
    with open(os.path.join(data_dir, "catalog.json"), "w") as f:
        json.dump(out, f, indent=1)

    for k, v in out.items():
        print(f"{k}: {len(v)}")


if __name__ == "__main__":
    main(sys.argv[1])
